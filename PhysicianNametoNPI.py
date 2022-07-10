# import module
import requests
import json
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from threading import Thread
import time

class PhysicianName:
  def __init__(self, first_name, last_name, state, row, second_last_name, npi):
    self.first_name = first_name
    self.last_name = last_name
    self.state = state
    self.row = row
    self.second_last_name = second_last_name
    self.npi = npi
# load excel with its path

#load excel to get list of fname, lname, state
def get_physician_name_from_excel(wrkbk, second_round):
    physician_names = []

    # load excel with its path

    sh = wrkbk['Temple']

    # iterate through excel and display data
    for i in range(1, sh.max_row+1):
        state = sh.cell(row = i, column = 4).value
        second_last_name = sh.cell(row = i, column = 3).value
        npi = sh.cell(row = i, column = 5).value
        physician_name = PhysicianName(sh.cell(row=i, column = 1).value, sh.cell(row=i, column=2).value, state, i, second_last_name, npi)
        physician_names.append(physician_name)
    get_npi_lists(physician_names, wrkbk, sh, second_round)
#process each entry in the list
def get_npi_lists(physician_names, wrkbk, sh, second_round):
    for physician_name in physician_names:
      if(physician_name.npi is None):
          t = Thread(target= get_npi_from_name, args=(physician_name, wrkbk, sh, second_round))
          t.start()
    wrkbk.save("/Users/ptran2/Downloads/ACOs_1.xlsx")

def get_npi_from_name(physician_name, wrkbk, sh, second_round):
    cell = 'E' + str(physician_name.row)
    #edge case multiple last names
    if(physician_name.second_last_name is None):
        if(physician_name.state is None):
            api_url = "https://npiregistry.cms.hhs.gov/api/?version=2.0&&pretty=true&first_name={fname}&last_name={lname}&use_first_name_alias=false".format(fname=physician_name.first_name, lname = physician_name.last_name)
        else:
            api_url = "https://npiregistry.cms.hhs.gov/api/?version=2.0&&pretty=true&state={st}&first_name={fname}&last_name={lname}&use_first_name_alias=false".format(st = physician_name.state,fname=physician_name.first_name, lname = physician_name.last_name)
        response = requests.get(api_url)
        result_count = response.json()["result_count"]
        #edge case if the doctor doesnt exist
        if(physician_name.row % 50 == 0):
            wrkbk.save("/Users/ptran2/Downloads/ACOs_1.xlsx")
            print(physician_name.row)
        if(result_count > 0):
            #cellref=sh.cell(row=physician_name.row, column=5)
            #cellref=response.json()["results"][0]["number"]
            address_cell = "F" + str(physician_name.row)
            taxonomy_cell = "G" + str(physician_name.row)
            if(result_count == 1):
                sh[cell] = response.json()["results"][0]["number"]
                address = response.json()["results"][0]["addresses"][0]
                sh[address_cell] = address["address_1"] + " " + address["address_2"] + ", " + address["city"] + ", " + address["state"] + ", " + address["postal_code"]
                taxonomies_list = response.json()["results"][0]["taxonomies"]
                for index in range(len(taxonomies_list)):
                    if(taxonomies_list[index]["primary"]):
                        sh[taxonomy_cell] = taxonomies_list[index]["desc"]
            else:
                npi_list = ""
                #edge case of multiple doctors in the same state with the same physician_name
                for i in range(0, result_count):
                    npi_list += str(response.json()["results"][i]["number"]) + ","
                sh[cell] = npi_list
                redFill = PatternFill(start_color='FFFF0000',
                                   end_color='FFFF0000',
                                   fill_type='solid')
                sh[cell].fill = redFill
        else:
            yellowFill = PatternFill(start_color='FFFF00',
                                           end_color='FFFF00',
                                           fill_type='solid')
            sh[cell].fill = yellowFill
            sh[cell] = "!"
    else:
        sh[cell] = "@"
        greenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        sh[cell].fill = greenFill
    if(second_round):
        wrkbk.save("/Users/ptran2/Downloads/ACOs_1.xlsx")
wrkbk = openpyxl.load_workbook("/Users/ptran2/Downloads/ACOs_1.xlsx")
get_physician_name_from_excel(wrkbk, False)
time.sleep(5)
get_physician_name_from_excel(wrkbk, True)




